import sys
import openpyxl
from openpyxl.utils import get_column_letter

#python testDataAnalyzer.py V32.xlsx BCP_CHIE01.txt 3 1 PROD_MWHE01.txt 3 3

def process_datafile(data_file):
    data_dict = {}
    with open(data_file, 'r') as f:
        lines = f.readlines()

    unit_name = ""
    unit_data = []
    for line in lines:
        line = line.strip()
        if line.startswith("-"):
            if unit_name and unit_data:
                data_dict[unit_name] = unit_data
                unit_name = ""
                unit_data = []
            continue

        if line:
            if not unit_name:
                unit_name = line
            else:
                unit_data.append(line)

    return data_dict

def fillin_excel(workbook, data_dict, row, col):
    try:
        tables_count_sheet = workbook["Tables Count"]
        template_sheet = workbook["Template"]

        for unit_name, unit_data in data_dict.items():
            for row_obj in tables_count_sheet.iter_rows(min_row=2, max_row=tables_count_sheet.max_row, min_col=1, max_col=1):
                if row_obj[0].value == unit_name:
                    matching_row = row_obj[0].row
                    unit_sheet_name = tables_count_sheet.cell(row=matching_row, column=2).value

                    if unit_sheet_name not in workbook.sheetnames:
                        #workbook.create_sheet(unit_sheet_name)
                        new_sheet = workbook.copy_worksheet(template_sheet)
                        new_sheet.title = unit_sheet_name

                    unit_sheet = workbook[unit_sheet_name]
                    for i, data_row in enumerate(unit_data, start=row):
                        data_values = data_row.split("\t")
                        for j, value in enumerate(data_values, start=col):
                            unit_sheet.cell(row=i, column=j, value=value)

                    print(f"Unit: {unit_name}, Sheet Name: {unit_sheet_name}, Row: {row}, Col: {get_column_letter(col)}")
                    break

    except Exception as e:
        print(f"Error: {e}")

def main():
    if len(sys.argv) != 8:
        print("Usage: python script.py excel_file data_file1 row1 col1 data_file2 row2 col2")
        return

    excel_file = sys.argv[1]
    data_file1 = sys.argv[2]
    row1 = int(sys.argv[3])
    col1 = int(sys.argv[4])
    data_file2 = sys.argv[5]
    row2 = int(sys.argv[6])
    col2 = int(sys.argv[7])

    try:
        workbook = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    print(f"Data File: {data_file1}")
    data_dict1 = process_datafile(data_file1)
    fillin_excel(workbook, data_dict1, row1, col1)

    print(f"Data File: {data_file2}")
    data_dict2 = process_datafile(data_file2)
    fillin_excel(workbook, data_dict2, row2, col2)

    workbook.save(excel_file)

if __name__ == "__main__":
    main()
