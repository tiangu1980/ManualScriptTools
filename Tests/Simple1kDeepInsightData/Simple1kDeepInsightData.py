import pandas as pd
from openpyxl import load_workbook  

output_sheet_to_overwrite='PivotData'
output_Excel_File='output_Name-MandM.xlsx'

source_excel_file = 'Data insights-OpenAI data Dec-23-0131.xlsx'
source_sheet_name = 'Speech 1k vs. Open AI paid'
source_columns = ['Cloud Customer GUID', 'Name', 'Current Status', 'Current CU', 'CU MoM', 'Account Manager', 'Country Name', 'Industry', 'Vertical', 'Segments', 'Current Month Meters', 'Kind', 'Month'] 

# 读取 Excel 文件数据
df = pd.read_excel(source_excel_file, sheet_name=source_sheet_name, usecols=source_columns)
df = df[(df['Current CU'] != 0) & df['Current CU'].notna()]
df['Month'] = df['Month'].dt.strftime("%Y_%m")

# 获取'Month'列的无重复值
print(df.columns)
months = df['Month'].unique().tolist()
months.sort()
print(months)
dfs_perMonth = []
for month in months:
    print(month)
    df_current = df[df['Month'] == month]
    
# 以 Month 和 Name 进行分组，并计算 Speech 和 AOAI 的总和
    grouped = df_current.groupby(['Month', 'Name', 'Kind'])['Current CU'].sum().reset_index()

# 使用 pivot_table 将数据重塑，Speech 和 AOAI 分别作为列
    pivot = grouped.pivot_table(index=['Month', 'Name'], columns='Kind', values='Current CU', aggfunc='sum').reset_index()

# 重命名列
    pivot.columns.name = None  # 清除列名
    pivot.columns = ['Month', 'Name', 'AOAI', 'Speech']

# 如果某个 Name 的 Speech 或 AOAI 为空，则填充为 0
    pivot.fillna(0, inplace=True)
    
    dfs_perMonth.append(pivot)

merged_df = pd.concat([df.set_index('Name') for df in dfs_perMonth], axis=1, join='outer')
merged_df = merged_df.reset_index()
print(merged_df.columns)
merged_df.fillna(0, inplace=True)

# 创建一个 ExcelWriter 对象，用于写入Excel文件  
with pd.ExcelWriter(output_Excel_File, engine='openpyxl', mode='a') as writer:    
    # 如果'output_Name-MandM.xlsx'文件已经存在，并且你想覆盖第一个sheet的内容  
    # 首先获取已存在的sheet名，以避免覆盖其他sheet  
    book = writer.book  
    sheet_names = book.sheetnames  
      
    # 如果存在，则删除原 output_sheet_to_overwrite  
    if output_sheet_to_overwrite in book.sheetnames:  
        idx = book.sheetnames.index(output_sheet_to_overwrite)  
        book.remove(book.worksheets[idx])  
    else:  
        print(f"Sheet {output_sheet_to_overwrite} does not exist in {output_Excel_File}.")  
        exit()
      
    # 将df写入到新的sheet中，sheet名为 output_sheet_to_overwrite 或原来第一个sheet的名字  
    merged_df.to_excel(writer, sheet_name=output_sheet_to_overwrite, index=False)  

print(f"Data imported successfully into {output_Excel_File}.")
