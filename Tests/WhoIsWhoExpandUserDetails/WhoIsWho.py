import pandas as pd
from openpyxl import load_workbook

#https://www.ces.tech/sessions-events/speaker-directory.aspx

def has_hyperlink(cell):
    return cell.hyperlink is not None

def print_first_column(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    key=""
    for row_number, row in enumerate(sheet.iter_rows(min_col=1, max_col=1), start=1):
        for cell in row:
            if has_hyperlink(cell):
                name = cell.value
                url = cell.hyperlink.display
                key = row_number
                print(f"Row {key}, name: {name}")
                print(f"Row {key}, url: {url}")
                data_dict[key] = [name]
                data_dict[key].append(url)
            else:
                data_dict[key].append(cell.value)

data_dict = {}
current_texts = []



def save_dict_to_excel(data_dict, excel_file_path):
    # 创建一个 DataFrame 以保存字典
    df = pd.DataFrame(list(data_dict.items()), columns=['Key', 'Value'])

    # 将每个字符串数组拆分为单独的列，假设默认为5个元素
    df[['Name', 'Link', 'Position', 'Company', 'Extra']] = pd.DataFrame(df['Value'].apply(lambda x: x + [''] * (5 - len(x))).tolist())

    # 删除原始的 'Value' 列
    df = df.drop('Value', axis=1)

    # 将 DataFrame 保存为 Excel 文件
    df.to_excel(excel_file_path, index=False)

if __name__ == "__main__":
    # Excel文件路径，根据实际文件路径修改
    excel_file_path = 'WhoBook1.xlsx'
    
    print_first_column(excel_file_path)
    print(data_dict)

    excel_file_path = 'WhoIsWho.xlsx'
    save_dict_to_excel(data_dict, excel_file_path)