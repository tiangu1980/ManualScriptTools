import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import sys

# 当对象excel中Month列从{MonthStart}开始temporary vertical 列没有值时，检索相关内容填写该值
# 建立四个字典：CCID2temporaryvertical， CCID2Vertical， CCID2Vertical， CCID2SubVertical 四个字典。
# 建立字典: Vertical2temporaryvertical， Industry2temporaryvertical， SubVertical2temporaryvertical。
# 为每一个没有 temporaryvertical 值的行的ccid， 做如下操作：
#    1. 查找 CCID2temporaryvertical 如果该 ccid 的 dict value 中 item 数量大于0。
#       a. 以 dict value 中出现次数最多的 key 值，作为该行的 temporaryvertical 值。
#       b. 否则，返回空字符串。
#    顺寻查找 CCID2Vertical， CCID2Vertical， CCID2SubVertical, 如果一个中没有这个 ccid 就找下一个。如果找到了：
#    2. 如果上个 step 的 temporaryvertical 值是空字符串：
#       a. 以 dict value 中出现次数最多的key值，作为 targetKey
#       c. 用 targetKey 查找 ****-temporaryvertical，出现次数最多的为 temporaryvertical。
#       d。 否则，返回空字符串。

#    3. 上述所有步骤都没有 temporaryvertical, 则以 step1 的备选项 key 为最终结果 temporaryvertical。

emptyValues=["nan", "unknown", "#n/a", "n/a"]
MonthStart = "2023/07"

def process_excel(df, colsrc, coldest):
    # 去除空格，将列名转为小写
    colsrc = colsrc.strip()#.lower()
    coldest = coldest.strip()#.lower()

    # 初始化字典
    result_dict = {}

    # 遍历整个 Excel
    for index, row in df.iterrows():
        # 获取当前行 colsrc 列的值
        key = str(row[colsrc])

        # 获取当前行 coldest 列的值
        value = str(row[coldest])

        # 跳过空值
        #if pd.notna(key) and pd.notna(value) and value.lower() not in emptyValues:
        if pd.notna(key):
            # 如果 key 不在字典中，初始化一个嵌套字典
            if key not in result_dict:
                result_dict[key] = {}
                
            if pd.notna(value) and value.lower() not in emptyValues:

                # 如果 value 不在嵌套字典中，初始化一个计数器
                if value not in result_dict[key]:
                    result_dict[key][value] = 0

                # 增加计数
                result_dict[key][value] += 1

    # 打印结果        
    #for key in sorted(result_dict.keys()):
    #    nested_dict = result_dict[key]
    #    print(f"{key}: {nested_dict}")
    return result_dict

def get_most_common_coldest(dictVal, colsrc_value):
    # 获取 colsrc_value 对应的 value，如果不存在则返回空字典
    colsrc_dict = dictVal.get(colsrc_value, {})

    # 如果 colsrc_dict 为空字典，返回空字符串
    if not colsrc_dict:
        return ""

    # 使用 max 函数找到 coldest 出现次数最高的键
    most_common_coldest = max(colsrc_dict, key=colsrc_dict.get)

    return most_common_coldest

def get_most_common_coldest_2(dictVal1, dictVal2, colsrc_value):
    # 获取 colsrc_value 对应的 value，如果不存在则返回空字典
    colsrc_dict = dictVal1.get(colsrc_value, {})

    # 如果 colsrc_dict 为空字典，返回空字符串
    if not colsrc_dict:
        return ""

    # 使用 max 函数找到 coldest 出现次数最高的键
    targetKey = max(colsrc_dict, key=colsrc_dict.get)
    coldest_dict = dictVal2.get(targetKey, {})
    
    # 如果 coldest_dict 为空字典，返回空字符串
    if not coldest_dict:
        return ""
        
    # 使用 max 函数找到 coldest 出现次数最高的键
    most_common_coldest = max(coldest_dict, key=coldest_dict.get)

    return most_common_coldest


if __name__ == "__main__":
    # 从命令行接收参数
    if len(sys.argv) < 2:
        print("Usage: python process_excel.py <excel_file_path> <colsrc> <coldest>")
        sys.exit(1)

    excel_file_path = sys.argv[1]

    # 读取 Excel 文件
    df = pd.read_excel(excel_file_path)
    
    # 调用处理函数
    CCID2temporaryvertical = process_excel(df, "Cloud Customer GUID", "temporaryvertical")
    CCID2Vertical = process_excel(df, "Cloud Customer GUID", "Vertical")
    CCID2Vertical = process_excel(df, "Cloud Customer GUID", "Industry")
    CCID2SubVertical = process_excel(df, "Cloud Customer GUID", "Sub_Vertical")
    Vertical2temporaryvertical = process_excel(df, "Vertical", "temporaryvertical")
    Industry2temporaryvertical = process_excel(df, "Industry", "temporaryvertical")
    SubVertical2temporaryvertical = process_excel(df, "Sub_Vertical", "temporaryvertical")
    
    print(f"CCID2temporaryvertical\t{len(CCID2temporaryvertical)}")
    print(f"CCID2Vertical\t{len(CCID2Vertical)}")
    print(f"CCID2Vertical\t{len(CCID2Vertical)}")
    print(f"CCID2SubVertical\t{len(CCID2SubVertical)}")
    print(f"Vertical2temporaryvertical\t{len(Vertical2temporaryvertical)}")
    print(f"Industry2temporaryvertical\t{len(Industry2temporaryvertical)}")
    print(f"SubVertical2temporaryvertical\t{len(SubVertical2temporaryvertical)}")
    
    dictPairInputs=[[CCID2Vertical, Vertical2temporaryvertical],[CCID2Vertical, Industry2temporaryvertical],[CCID2SubVertical, SubVertical2temporaryvertical]]
    
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    sh = wb[wb.sheetnames[0]]
    
    # 获取列索引
    month_col_index = None
    temp_vertical_col_index = None
    CCID_col_index = None
    
    # 遍历第一行，找到 "Month" 和 "temporaryvertical" 列的索引
    for col_index, cell in enumerate(sh[1], 1):
        if cell.value == "Month":
            month_col_index = col_index
        elif cell.value == "temporaryvertical":
            temp_vertical_col_index = col_index
        elif cell.value == "Cloud Customer GUID":
            CCID_col_index = col_index
    
    # 如果找到了相应的列索引
    if month_col_index is not None and temp_vertical_col_index is not None:
        # 遍历每一行，更新符合条件的数据
        for row_number, row in enumerate(sh.iter_rows(min_row=2, values_only=True), start=2):
            month_value = row[month_col_index - 1]
            temp_vertical_value = row[temp_vertical_col_index - 1]
            CCID_value = row[CCID_col_index - 1]
    
            if month_value and temp_vertical_value is None and month_value >= MonthStart:
                # 更新 temporaryvertical 列的值为 "Test"
                #sh.cell(row=row_number, column=temp_vertical_col_index, value="Test")
    
                temporaryverticalVal = get_most_common_coldest(CCID2temporaryvertical, CCID_value)
                if not temporaryverticalVal:
                    for dictPair in dictPairInputs:
                        temporaryverticalVal = get_most_common_coldest_2(dictPair[0], dictPair[1], CCID_value)
                        if temporaryverticalVal:
                            break
                if temporaryverticalVal:
                    sh.cell(row=row_number, column=temp_vertical_col_index, value=temporaryverticalVal)
                    print(f"{row_number}\t{CCID_value}\t{month_value}\t{temporaryverticalVal}")
                else:
                    print(f"{row_number}\t{month_value}")
    
    # 保存工作簿到原文件
    wb.save(excel_file_path)