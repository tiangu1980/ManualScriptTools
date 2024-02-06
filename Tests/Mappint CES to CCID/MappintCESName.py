import re
from openpyxl import load_workbook

def clean_string(value):
    # 将字符串小写化，去除非字母字符，去除头尾空格，将连续空格转换为一个空格
    pattern = re.compile(r'(?<=[A-Z])\.(?=[A-Z])')
    result = re.sub(pattern, '', value)
    pattern = re.compile(r'(?<=[A-Z])\&(?=[A-Z])')
    result = re.sub(pattern, '', result)
    cleaned_value = re.sub(r'[^0-9a-zA-Z]', ' ', result.lower().strip())
    cleaned_value = re.sub(r'\bit\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\band\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bgmbh\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bltd\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\blimited\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bllc\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\binc\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bco\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bcompany\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bcorporation\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bcorp\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bbusiness\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bacademy\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bsoftware\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bapplication\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bapplications\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bproject\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bprojects\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bscience\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bnetwork\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bnetworks\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bretail\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bretails\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bsales\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bautomotive\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\btrans\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\btransp\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\btransportation\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bdigital\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bdept\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bdepartment\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bservice\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bservices\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\blab\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\blabs\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\blaboratory\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\btech\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\btechnical\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\btechnology\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\btechnologies\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bfactory\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\boffice\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bgroup\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bgroupe\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bintl\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\binternational\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bcom\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bgame\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bgames\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bgaming\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bmedia\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bmedias\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\buniv\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\buniversity\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\binfo\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\binformation\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bdevice\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bdevices\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bequip\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bequipment\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bequipments\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bdata\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bstudio\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bsys\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bsystem\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bsystems\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\benterprise\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\benterprises\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bcomm\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bcommunication\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bcommunications\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bfdn\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bfound\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bfoundation\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bfoundations\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bvideo\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bvideos\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\belectronic\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\belectronics\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bresearch\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bcentre\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bcenter\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bindustry\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bindustrial\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\btrading\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bsolution\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bsolutions\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bentertainment\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\bentertainments\b', ' ', cleaned_value)
    cleaned_value = re.sub(r'\s+', ' ', cleaned_value)
    words = cleaned_value.split()
    filtered_words = [word for word in words if len(word) > 1]
    result_string = ' '.join(filtered_words)
    return result_string
    
def compare_strings(str1, str2):
    #print(f"compare_strings({str1}, {str2})")
    # 将字符串用空格分隔为字符串数组
    strs1 = str1.split()
    strs2 = str2.split()
    
    rate1=0
    rate2=0

    # 初始化匹配计数
    match_count = 0

    # 如果 strs2 不为空，遍历 strs2，检查其是否存在于 strs1 中
    if strs2:
        for word in strs2:
            if word in strs1:
                match_count += 1

        # 计算匹配比例
        rate1 = match_count / len(strs2)
    else:
        # 处理 strs2 为空的情况，避免除以零错误
        rate1 = 0.00
        
    # 初始化匹配计数
    match_count = 0

    # 如果 strs1 不为空，遍历 strs1，检查其是否存在于 strs2 中
    if strs1:
        for word in strs1:
            if word in strs2:
                match_count += 1

        # 计算匹配比例
        rate2 = match_count / len(strs1)
    else:
        # 处理 strs2 为空的情况，避免除以零错误
        rate2 = 0.00
    
    return (rate1 + rate2)/2

# 读取文件1
workbook_file1 = load_workbook('ALL Speech Paid users-202312.xlsx')
sheet_file1 = workbook_file1.active

# 在sheet_file1末尾添加"ModifiedName"列
sheet_file1.cell(row=1, column=13, value="ModifiedName")

# 处理每一行
for row in range(2, sheet_file1.max_row + 1):
    en_name_value = sheet_file1.cell(row=row, column=3).value
    modified_name_value = clean_string(en_name_value)
    sheet_file1.cell(row=row, column=13, value=modified_name_value)
    #print(sheet_file1.cell(row=row, column=13).value)

# 在sheet_file1末尾添加5列
#for col_num, col_name in enumerate(["Logo", "Exhibitor", "Summary", "Booth", "FEATURED EXHIBITORS"], start=14):
#    sheet_file1.cell(row=1, column=col_num, value=col_name)

# 读取文件2
workbook_file2 = load_workbook('CES-Customers.xlsx')
sheet_file2 = workbook_file2.active

# 在sheet_file2末尾添加"ModifiedExhibitor"列
sheet_file2.cell(row=1, column=6, value="ModifiedExhibitor")

# 处理每一行
for row in range(2, sheet_file2.max_row + 1):
    exhibitor_value = sheet_file2.cell(row=row, column=2).value
    modified_exhibitor_value = clean_string(exhibitor_value)
    sheet_file2.cell(row=row, column=6, value=modified_exhibitor_value)
    #print(sheet_file2.cell(row=row, column=6).value)


# 遍历sheet_file1的“ModifiedName”列
for row in range(2, sheet_file1.max_row + 1):
    cur_modified_name = sheet_file1.cell(row=row, column=13).value

    dict_match={}
    
    # 遍历sheet_file2的“ModifiedExhibitor”列
    for row2 in range(2, sheet_file2.max_row + 1):
        cur_modified_exhibitor = sheet_file2.cell(row=row2, column=6).value
        
        # 保存列表
        tar_row2=row2
        cur_match_rate=compare_strings(cur_modified_exhibitor, cur_modified_name)
        if (cur_match_rate not in dict_match) and (cur_match_rate>0.00):
            dict_match[tar_row2] = cur_match_rate

        #print(f"{row} {cur_modified_name} {row2} {cur_modified_exhibitor} {cur_match_rate} {len(dict_match)} {dict_match[cur_match_rate]}")

        # 如果匹配，将对应行的内容复制到sheet_file1
        #if cur_modified_name == cur_modified_exhibitor:
        #    print(f"Match found at sheet_file1 row {row} and sheet_file2 row {row2}")
        #    for col_num, col_name in enumerate(["Logo", "Exhibitor", "Summary", "Booth", "FEATURED EXHIBITORS"], start=14):
        #        sheet_file1.cell(row=row, column=col_num, value=sheet_file2.cell(row=row2, column=col_num - 8).value)
        
    if dict_match:
        max_rate = max(dict_match.values())
        max_rows = [key for key, value in dict_match.items() if value == max_rate]
        matches=len(max_rows)
            # 如果 max_rows 小于 2，取出 dict_match 中 value 次高的 key 补充进 max_rows
        if matches < 2 and len(dict_match)>1:
            remaining_keys = [key for key, value in dict_match.items() if key not in max_rows]
            remaining_keys_sorted = sorted(remaining_keys, key=lambda k: dict_match[k], reverse=True)
        
            # 取出次高的 key 值
            additional_keys = remaining_keys_sorted[:2 - matches]
        
            # 将补充的 key 值添加到 max_rows 中
            max_rows.extend(additional_keys)
        matches=len(max_rows)
        #for key in max_keys:
        #    print(f"Row1: {row}  LenDict: {matches}  Rate: {max_rate}  Row2: {Key}")
        #    sheet_file1.cell(row=row, column=col_num, value=sheet_file2.cell(row=match_row2, column=col_num - 8).value)    
        tcolumn = 14
        for key in max_rows:
            namef2=sheet_file2.cell(row=key, column=2).value
            namef2b=sheet_file2.cell(row=key, column=6).value
            print(f"Row1: {row}  LenDict: {matches}  Rate: {max_rate}  Row2: {key}  Value2: {namef2}  Value2b: {namef2b}")
            
            # 将 key 填写到当前格子
            sheet_file1.cell(row=row, column=tcolumn, value=dict_match[key])
            sheet_file1.cell(row=row, column=tcolumn+1, value=key)
            sheet_file1.cell(row=row, column=tcolumn+2, value=namef2)
            sheet_file1.cell(row=row, column=tcolumn+3, value=namef2b)
            
            # 列数加 1
            tcolumn += 4   
    #input(ss)
# 保存文件1
workbook_file1.save('output_file1.xlsx')
# 保存文件2
workbook_file2.save('output_file2.xlsx')
