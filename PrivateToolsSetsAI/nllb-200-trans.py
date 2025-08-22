import os
import time
import argparse
import shutil
from openpyxl import load_workbook
from transformers import AutoTokenizer, AutoModelForSeq2SeqLM

# ========== 语言映射表 ==========
LANG_MAP = {
    # 常见欧洲语言
    "en": "eng_Latn", "en-us": "eng_Latn", "en-gb": "eng_Latn",
    "cs": "ces_Latn", "cs-cz": "ces_Latn",
    "da": "dan_Latn", "da-dk": "dan_Latn",
    "de": "deu_Latn", "de-de": "deu_Latn", "de-at": "deu_Latn", "de-ch": "deu_Latn",
    "es": "spa_Latn", "es-es": "spa_Latn", "es-mx": "spa_Latn", "es-ar": "spa_Latn",
    "fr": "fra_Latn", "fr-fr": "fra_Latn", "fr-ca": "fra_Latn", "fr-ch": "fra_Latn",
    "it": "ita_Latn", "it-it": "ita_Latn",
    "nl": "nld_Latn", "nl-nl": "nld_Latn", "nl-be": "nld_Latn",
    "pl": "pol_Latn", "pl-pl": "pol_Latn",
    "pt": "por_Latn", "pt-pt": "por_Latn", "pt-br": "por_Latn",
    "ru": "rus_Cyrl", "ru-ru": "rus_Cyrl",
    "uk": "ukr_Cyrl", "uk-ua": "ukr_Cyrl",

    # 北欧语言
    "sv": "swe_Latn", "sv-se": "swe_Latn",
    "no": "nob_Latn", "nb": "nob_Latn", "nb-no": "nob_Latn", "nn-no": "nno_Latn",
    "fi": "fin_Latn", "fi-fi": "fin_Latn",
    "is": "isl_Latn", "is-is": "isl_Latn",

    # 东亚语言
    "zh": "zho_Hans", "zh-cn": "zho_Hans", "zh-sg": "zho_Hans",
    "zh-hans": "zho_Hans", "zh-hant": "zho_Hant",
    "zh-tw": "zho_Hant", "zh-hk": "zho_Hant", "zh-mo": "zho_Hant",
    "ja": "jpn_Jpan", "ja-jp": "jpn_Jpan",
    "ko": "kor_Hang", "ko-kr": "kor_Hang",

    # 中东语言
    "ar": "arb_Arab", "ar-sa": "arb_Arab", "ar-eg": "arb_Arab", "ar-dz": "arb_Arab",
    "ar-ma": "arb_Arab", "ar-tn": "arb_Arab", "ar-jo": "arb_Arab", "ar-iq": "arb_Arab",
    "ar-ly": "arb_Arab", "ar-sy": "arb_Arab", "ar-ye": "arb_Arab",
    "he": "heb_Hebr", "he-il": "heb_Hebr", "iw": "heb_Hebr",

    # 南亚语言
    "hi": "hin_Deva", "hi-in": "hin_Deva",
    "bn": "ben_Beng", "bn-bd": "ben_Beng", "bn-in": "ben_Beng",
    "ur": "urd_Arab", "ur-pk": "urd_Arab", "ur-in": "urd_Arab",
    "ta": "tam_Taml", "ta-in": "tam_Taml", "ta-lk": "tam_Taml",
    "te": "tel_Telu", "te-in": "tel_Telu",
    "ml": "mal_Mlym", "ml-in": "mal_Mlym",
    "mr": "mar_Deva", "mr-in": "mar_Deva",
    "gu": "guj_Gujr", "gu-in": "guj_Gujr",

    # 东南亚语言
    "th": "tha_Thai", "th-th": "tha_Thai",
    "vi": "vie_Latn", "vi-vn": "vie_Latn",
    "id": "ind_Latn", "id-id": "ind_Latn",
    "ms": "zsm_Latn", "ms-my": "zsm_Latn",

    # 非洲语言（部分）
    "sw": "swh_Latn", "sw-ke": "swh_Latn", "sw-tz": "swh_Latn",
    "am": "amh_Ethi", "am-et": "amh_Ethi",
    "zu": "zul_Latn", "zu-za": "zul_Latn",
    "xh": "xho_Latn", "xh-za": "xho_Latn",
    "st": "sot_Latn", "st-za": "sot_Latn",

    # 其他常见语言
    "tr": "tur_Latn", "tr-tr": "tur_Latn",
    "fa": "pes_Arab", "fa-ir": "pes_Arab",
    "el": "ell_Grek", "el-gr": "ell_Grek",
    "hu": "hun_Latn", "hu-hu": "hun_Latn",
    "ro": "ron_Latn", "ro-ro": "ron_Latn",
    "bg": "bul_Cyrl", "bg-bg": "bul_Cyrl",
    "sr": "srp_Cyrl", "sr-rs": "srp_Cyrl",
    "hr": "hrv_Latn", "hr-hr": "hrv_Latn",
    "sk": "slk_Latn", "sk-sk": "slk_Latn",
    "sl": "slv_Latn", "sl-si": "slv_Latn",
}

#def normalize_lang_code(lang_code):
#    """将 cs-cz / zh_cn 这种格式转换为 NLLB-200 语言代码"""
#    if not lang_code:
#        return None
#    lc = lang_code.lower().replace("_", "-")
#    return LANG_MAP.get(lc, None)

def normalize_lang_code(market_code: str) -> str:
    """
    将 Market 值（如 zh-cn, en-us, sr-latn-rs）映射为 NLLB-200 语言代码
    """
    market_code = market_code.lower().replace("_", "-")
    
    # 直接命中
    if market_code in LANG_MAP:
        return LANG_MAP[market_code]
    
    # 只取前缀（语言部分）
    base = market_code.split("-")[0]
    if base in LANG_MAP:
        return LANG_MAP[base]
    
    print(f"⚠️ 未知映射 {market_code}，使用英语 fallback")
    return "eng_Latn"
    

# ========== 加载 NLLB-200 模型 ==========
#MODEL_PATH = "./models/nllb-200-distilled-600M"
MODEL_PATH = r"D:\GrowthAI\models\nllb-200-distilled-600M"
tokenizer = AutoTokenizer.from_pretrained(MODEL_PATH)
model = AutoModelForSeq2SeqLM.from_pretrained(MODEL_PATH)

def translate_text(text, src_lang, tgt_lang, max_length=512):
    if not text:
        return ""
    try:
        inputs = tokenizer(text, return_tensors="pt", padding=True, truncation=True, max_length=max_length)
        translated_tokens = model.generate(
            **inputs,
            forced_bos_token_id=tokenizer.lang_code_to_id[tgt_lang],  # 目标语言
            max_length=max_length
        )
        return tokenizer.batch_decode(translated_tokens, skip_special_tokens=True)[0]
    except Exception as e:
        print(f"翻译出错: {e}")
        return ""

def process_excel(input_path, output_path, column_name, src_lang, tgt_lang):
    # 创建备份
    backup_path = input_path.replace('.xlsx', '_backup.xlsx')
    shutil.copy2(input_path, backup_path)
    print(f"已创建备份: {backup_path}")

    # 读取 Excel
    wb = load_workbook(input_path)
    ws = wb.active

    # 找到列索引
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        raise ValueError(f"找不到列: {column_name}")
    col_idx = headers.index(column_name) + 1

    # 如果 src_lang 未指定，则必须有 "Market" 列
    market_idx = None
    if not src_lang:
        if "Market" not in headers:
            raise ValueError("未指定 --src_lang 且 Excel 中没有 Market 列")
        market_idx = headers.index("Market") + 1

    # 添加新列标题
    new_col_name = f"Feedback_{tgt_lang}"
    ws.cell(row=1, column=ws.max_column + 1, value=new_col_name)
    new_col_idx = ws.max_column

    # 逐行翻译
    for row_idx in range(2, ws.max_row + 1):
        text = ws.cell(row=row_idx, column=col_idx).value
        if not text:
            continue

        # 动态获取源语言
        row_src_lang = src_lang
        if not row_src_lang and market_idx:
            market_val = ws.cell(row=row_idx, column=market_idx).value
            row_src_lang = normalize_lang_code(market_val)
        else:
            row_src_lang = normalize_lang_code(row_src_lang)

        row_tgt_lang = normalize_lang_code(tgt_lang)

        if not row_src_lang or not row_tgt_lang:
            print(f"⚠️ 第{row_idx}行无法识别语言代码，跳过 (src={src_lang}, tgt={tgt_lang})")
            continue

        print(f"正在处理第{row_idx}行: {text} ({row_src_lang} → {row_tgt_lang})")
        translated = translate_text(text, row_src_lang, row_tgt_lang)
        ws.cell(row=row_idx, column=new_col_idx, value=translated)
        print(f"已翻译: {translated}")
        time.sleep(0.2)

    # 保存结果
    wb.save(output_path)
    print(f"✅ 翻译完成，结果已保存到: {output_path}")

def main():
    parser = argparse.ArgumentParser(description="Excel翻译工具 (NLLB-200)")
    parser.add_argument("--input", required=True, help="输入Excel文件路径 (xlsx)")
    parser.add_argument("--output", required=True, help="输出Excel文件路径 (xlsx)")
    parser.add_argument("--column", required=True, help="要翻译的列名")
    parser.add_argument("--src_lang", required=False, help="源语言代码 (可用 cs-cz, zh_cn 等格式，也可留空自动从 Market 列获取)")
    parser.add_argument("--tgt_lang", required=True, help="目标语言代码 (可用 en-us, ja-jp, de-de 等格式)")

    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"⚠️ 输入文件不存在: {args.input}")
        return

    try:
        process_excel(args.input, args.output, args.column, args.src_lang, args.tgt_lang)
    except Exception as e:
        print(f"❌ 处理失败: {str(e)}")

if __name__ == "__main__":
    main()
