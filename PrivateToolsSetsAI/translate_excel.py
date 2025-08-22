import os
import time
import argparse
import shutil
from googletrans import Translator
from openpyxl import load_workbook

def translate_text(text, dest_lang, max_retries=3, retry_delay=1):
    """翻译文本，带重试机制"""
    for attempt in range(max_retries + 1):
        try:
            translator = Translator(service_urls=['translate.google.com'])
            result = translator.translate(text, dest=dest_lang)
            return result.text
        except Exception as e:
            if attempt < max_retries:
                print(f"第{attempt+1}次尝试失败，{retry_delay}秒后重试...")
                time.sleep(retry_delay)
            else:
                print(f"翻译到{dest_lang}时出错（已达最大重试次数）: {e}")
                return ""

def process_excel(input_path, output_path, column_name, target_lang):
    # 备份输入文件
    backup_path = input_path.replace('.xlsx', '_backup.xlsx')
    shutil.copy2(input_path, backup_path)
    print(f"已创建备份: {backup_path}")

    # 读取 Excel
    wb = load_workbook(input_path)
    ws = wb.active

    # 找到列索引
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        raise ValueError(f"找不到列: {column_name}")
    col_idx = headers.index(column_name) + 1

    # 添加新列标题
    new_col_name = f"Feedback_{target_lang}"
    ws.cell(row=1, column=ws.max_column + 1, value=new_col_name)
    new_col_idx = ws.max_column

    # 逐行翻译
    for row_idx in range(2, ws.max_row + 1):
        text = ws.cell(row=row_idx, column=col_idx).value
        if text:
            print(f"正在处理第{row_idx}行: {text}")
            translated = translate_text(text, target_lang)
            ws.cell(row=row_idx, column=new_col_idx, value=translated)
            print(f"已翻译为{target_lang}: {translated}")
            time.sleep(1)  # 避免过快请求

    # 保存结果
    wb.save(output_path)
    print(f"✅ 翻译完成，结果已保存到: {output_path}")

def main():
    parser = argparse.ArgumentParser(description="Excel翻译工具")
    parser.add_argument("--input", required=True, help="输入Excel文件路径 (xlsx)")
    parser.add_argument("--output", required=True, help="输出Excel文件路径 (xlsx)")
    parser.add_argument("--column", required=True, help="要翻译的列名")
    parser.add_argument("--lang", required=True, help="目标语言代码，如 en, ja, fr, de 等")

    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"⚠️ 输入文件不存在: {args.input}")
        return

    try:
        process_excel(args.input, args.output, args.column, args.lang)
    except Exception as e:
        print(f"❌ 处理失败: {str(e)}")

if __name__ == "__main__":
    main()
